"""
etl/excel_reader.py — Lee cartera.xlsx (QRY de cartera) y calcula la composición.

Formato esperado: hoja 'Consulta1' con columnas crudas (una fila por instrumento):
  DSC_CUENTA (fondo), COD_MONEDA, TRAMO, CLASIFICACION_OBS, VALOR_PRESENTE_MERCADO.
La composición (moneda / duración / instrumento) se agrega aquí ponderando por
VALOR_PRESENTE_MERCADO. Si el archivo no trae 'Consulta1' pero sí las hojas
pre-agregadas ('moneda'/'duracion'/'instrumento'), se usa el lector antiguo (fallback).
"""
import unicodedata
import openpyxl
from config import ARCHIVO_CARTERA

# Tramos de duración en el orden del folleto (incluye los de 0% para que se muestren todos)
ORDEN_TRAMOS = ["Hasta 30 días", "Entre 31 y 90 días", "Entre 91 y 120 días", "Entre 1 y 2 años"]


def _norm(s: str) -> str:
    """Normaliza para comparar nombres de fondo: sin acentos, mayúsculas, espacios colapsados."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return " ".join(s.upper().split())


def _fmt(v: float) -> str:
    return f"{v*100:.2f}%".replace(".", ",")


def _hdr_cols(ws) -> dict:
    """Mapea nombre de encabezado (fila 1) -> índice 0-based."""
    cols = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is not None:
            cols[str(h).strip().upper()] = c - 1
    return cols


def _from_consulta1(ws, nombre_fondo: str) -> dict:
    """Agrega la composición del fondo desde la hoja cruda Consulta1."""
    cols = _hdr_cols(ws)
    c_fund  = cols.get("DSC_CUENTA")
    c_mon   = cols.get("COD_MONEDA")
    c_tramo = cols.get("TRAMO")
    c_clas  = cols.get("CLASIFICACION_OBS")
    c_val   = cols.get("VALOR_PRESENTE_MERCADO")
    if None in (c_fund, c_mon, c_tramo, c_clas, c_val):
        return {"moneda": [], "duracion": [], "instrumento": []}

    target = _norm(nombre_fondo)
    mon_sum, dur_sum, inst_sum = {}, {}, {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        fund = row[c_fund] if c_fund < len(row) else None
        if not fund:
            continue
        if _norm(fund) != target:   # match exacto normalizado (sin acentos)
            continue
        val = row[c_val] if c_val < len(row) else None
        if not isinstance(val, (int, float)):
            continue
        mon   = row[c_mon]   if c_mon   < len(row) else None
        tramo = row[c_tramo] if c_tramo < len(row) else None
        clas  = row[c_clas]  if c_clas  < len(row) else None
        if mon:   mon_sum[str(mon).strip()]    = mon_sum.get(str(mon).strip(), 0.0)   + val
        if tramo: dur_sum[str(tramo).strip()]  = dur_sum.get(str(tramo).strip(), 0.0) + val
        if clas:  inst_sum[str(clas).strip()]  = inst_sum.get(str(clas).strip(), 0.0) + val

    result = {"moneda": [], "duracion": [], "instrumento": []}
    total = sum(mon_sum.values())
    if total > 0:
        # MONEDA: sólo > 0.1%, ordenado desc
        for mon, v in sorted(mon_sum.items(), key=lambda x: -x[1]):
            if v / total > 0.001:
                result["moneda"].append((mon, _fmt(v / total)))
        # DURACIÓN: los 4 tramos en orden (incluye 0% para que se muestren todos)
        for tramo in ORDEN_TRAMOS:
            result["duracion"].append((tramo, _fmt(dur_sum.get(tramo, 0.0) / total)))
        # INSTRUMENTO: sólo > 0.1%, ordenado desc
        for inst, v in sorted(inst_sum.items(), key=lambda x: -x[1]):
            if v / total > 0.001:
                result["instrumento"].append((inst, _fmt(v / total)))
    return result


def get_cartera_composicion(nombre_fondo: str) -> dict:
    """Lee composición (moneda, duración, instrumento) desde cartera.xlsx."""
    result = {"moneda": [], "duracion": [], "instrumento": []}
    try:
        wb = openpyxl.load_workbook(str(ARCHIVO_CARTERA), read_only=True, data_only=True)
    except Exception as e:
        print(f"    [WARN] No se pudo leer cartera.xlsx: {e}")
        return result

    try:
        if "Consulta1" in wb.sheetnames:
            result = _from_consulta1(wb["Consulta1"], nombre_fondo)
        else:
            print("    [WARN] cartera.xlsx sin hoja 'Consulta1'")
    except Exception as e:
        print(f"    [WARN] Error agregando composición: {e}")
    finally:
        wb.close()
    return result
