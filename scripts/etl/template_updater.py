"""
etl/template_updater.py — Actualiza todos los templates con datos del mes nuevo.

FLUJO:
1. Determina el mes objetivo
2. Obtiene VC de fin de mes para cada fondo desde:
   a) inputs/planilla_vc.xlsx (si el usuario la subió)
   b) ODS VALORES_CUOTA_GPI (si tiene día >= 28 del mes)
3. Obtiene ICP y Santander MM del mes
4. Para cada template, usa DETECCIÓN DE FÓRMULAS para encontrar:
   - Qué fila tiene los VC de ICP/FIP/Comp para el año actual
   - Qué columna es la siguiente a llenar (mayo = col T si enero fue P)
   - Qué fila tiene los retornos mensuales
   - Qué fila tiene los totales anuales
5. Escribe los nuevos datos y fórmulas en las celdas correctas
6. Actualiza AW19 con la nueva fecha de cierre
"""

import os, calendar, requests, json, re
from datetime import date, datetime
from pathlib import Path
from dateutil.relativedelta import relativedelta

import openpyxl
import openpyxl.utils as xl_u

_INPUTS   = Path(__file__).parent.parent.parent / "inputs"
_TMPL_DIR = _INPUTS / "templates"
_ODS_API  = "https://claudeods.vantrustcapital.cl/query"

FUND_ODS_NAMES = {
    'TEMPLATE FONDO LIQUIDEZ ACTIVA.xlsx':        'FIP VANTRUST LIQUIDEZ ACTIVA',
    'TEMPLATE FONDO ALTO APORTE.xlsx':            'FIP VANTRUST LIQUIDEZ ALTO APORTE',
    'TEMPLATE FONDO ALTO CAPITAL.xlsx':           'FIP VANTRUST LIQUIDEZ ALTO CAPITAL',
    'TEMPLATE FONDO LIQUIDEZ ALTO MONTO.xlsx':    'FIP VANTRUST LIQUIDEZ ALTO MONTO',
    'TEMPLATE FONDO LIQUIDEZ CAJA.xlsx':          'FIP VANTRUST LIQUIDEZ CAJA',
    'TEMPLATE FONDO LIQUIDEZ CONTINUA.xlsx':      'FIP VANTRUST LIQUIDEZ CONTINUA',
    'TEMPLATE FONDO LIQUIDEZ CORRIENTE.xlsx':     'FIP VANTRUST LIQUIDEZ CORRIENTE',
    'TEMPLATE FONDO LIQUIDEZ CORTO PLAZO.xlsx':   'FIP VANTRUST LIQUIDEZ CORTO PLAZO',
    'TEMPLATE FONDO LIQUIDEZ Disponible I.xlsx':  'FIP VANTRUST LIQUIDEZ DISPONIBLE I',
    'TEMPLATE FONDO LIQUIDEZ DOLAR.xlsx':         'FIP VANTRUST LIQUIDEZ DOLAR',
    'TEMPLATE FONDO LIQUIDEZ DOLAR CAJA.xlsx':    'FIP VANTRUST LIQUIDEZ DOLAR CAJA',
    'TEMPLATE FONDO LIQUIDEZ EFECTIVO.xlsx':      'FIP VANTRUST LIQUIDEZ EFECTIVO',
    'TEMPLATE FONDO LIQUIDEZ FLEXIBLE.xlsx':      'FIP VANTRUST LIQUIDEZ FLEXIBLE',
    'TEMPLATE FONDO LIQUIDEZ UNO.xlsx':           'FIP VANTRUST LIQUIDEZ I',
    'TEMPLATE FONDO LIQUIDEZ LOCAL.xlsx':         'FIP VANTRUST LIQUIDEZ LOCAL',
    'TEMPLATE FONDO LIQUIDEZ Monetario I.xlsx':   'FIP VANTRUST LIQUIDEZ MONETARIO I',
    'TEMPLATE FONDO LIQUIDEZ Permanente.xlsx':    'FIP VANTRUST LIQUIDEZ PERMANENTE',
    'TEMPLATE FONDO LIQUIDEZ PLUS.xlsx':          'FIP VANTRUST LIQUIDEZ PLUS',
    'TEMPLATE FONDO LIQUIDEZ Presente.xlsx':      'FIP VANTRUST LIQUIDEZ PRESENTE',
    'TEMPLATE FONDO LIQUIDEZ RENDIMIENTO.xlsx':   'FIP VANTRUST LIQUIDEZ RENDIMIENTO',
    'TEMPLATE FONDO LIQUIDEZ RESERVA DOLAR.xlsx': 'FIP VANTRUST LIQUIDEZ RESERVA DÓLAR',
    'TEMPLATE FONDO LIQUIDEZ SENCILLO.xlsx':      'FIP VANTRUST LIQUIDEZ SENCILLO',
}

def _eom(y, m):
    return date(y, m, calendar.monthrange(y, m)[1])

# ── Data fetching ─────────────────────────────────────────────────────────────
def _ods_vc(nombre: str, y: int, m: int):
    sql = (f"SELECT MAX(VALOR_CUOTA) vc, MAX(DAY(FECHA_CIERRE)) last_day "
           f"FROM ODS.VALORES_CUOTA_GPI "
           f"WHERE RTRIM(LTRIM(EMPRESA))=N'{nombre}' AND VALOR_CUOTA>0 "
           f"AND YEAR(FECHA_CIERRE)={y} AND MONTH(FECHA_CIERRE)={m}")
    try:
        r = requests.post(_ODS_API, json={"Sql": sql},
                          headers={"Content-Type": "application/json"}, timeout=30)
        row = r.json().get("rows", [{}])[0]
        if row.get("last_day") and int(row["last_day"]) >= 28:
            return float(row["vc"])
        if row.get("last_day"):
            print(f"  [WARN] ODS {nombre} {y}-{m:02d}: only day {row['last_day']} (not end-of-month)")
    except Exception as e:
        print(f"  [WARN] ODS failed for {nombre}: {e}")
    return None

def _get_icp(y: int, m: int):
    try:
        with open(_INPUTS / "icp_clicp.json") as f:
            d = json.load(f)
        return float(d.get(f"{y}-{m:02d}") or d.get(f"{y}-{m}") or 0) or None
    except: return None

def _get_santander(y: int, m: int):
    try:
        with open(_INPUTS / "comp_clp.json") as f:
            d = json.load(f)
        return float(d.get(f"{y}-{m:02d}") or d.get(f"{y}-{m}") or 0) or None
    except: return None

def _read_planilla(y: int, m: int) -> dict:
    """
    Lee planilla_vc.xlsx con el output del query PUBLICADOR_PRECIO.
    Formato esperado: columnas FECHA, NEMOTECNICO, PRECIO (y otras opcionales).
    Devuelve {nemotecnico: precio} con el último precio del mes y/m para cada fondo.
    """
    path = _INPUTS / "planilla_vc.xlsx"
    if not path.exists():
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        # Find header row: look for FECHA, NEMOTECNICO, PRECIO columns
        fecha_col = nemo_col = precio_col = None
        header_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                v = str(ws.cell(r, c).value or '').strip().upper()
                if v == 'FECHA':    fecha_col  = c; header_row = r
                if v in ('NEMOTECNICO', 'NEMO', 'FONDO', 'FUND'): nemo_col = c; header_row = r
                if v in ('PRECIO', 'VC', 'VALOR_CUOTA', 'VALOR CUOTA'): precio_col = c; header_row = r
            if fecha_col and nemo_col and precio_col:
                break

        if not (fecha_col and nemo_col and precio_col):
            print(f"  [WARN] planilla_vc.xlsx: no se encontraron columnas FECHA/NEMOTECNICO/PRECIO")
            print(f"  Columnas detectadas: {[ws.cell(1,c).value for c in range(1,ws.max_column+1)]}")
            wb.close(); return {}

        # Read rows: collect last price per nemo for the target month
        vc_map = {}  # nemo -> (fecha, precio) keeping the latest date
        for r in range(header_row + 1, ws.max_row + 1):
            fecha  = ws.cell(r, fecha_col).value
            nemo   = ws.cell(r, nemo_col).value
            precio = ws.cell(r, precio_col).value

            if not (fecha and nemo and precio): continue
            if not isinstance(precio, (int, float)) or precio <= 0: continue
            if not hasattr(fecha, 'year'): continue
            if fecha.year != y or fecha.month != m: continue

            nemo_clean = str(nemo).strip()
            existing = vc_map.get(nemo_clean)
            if existing is None or fecha > existing[0]:
                vc_map[nemo_clean] = (fecha, float(precio))

        wb.close()
        result = {k: v[1] for k, v in vc_map.items()}
        if result:
            print(f"  Planilla: {len(result)} fondos para {y}-{m:02d}")
            # Show a few for verification
            for k, v in list(result.items())[:3]:
                print(f"    {k}: {v}")
        else:
            print(f"  [WARN] planilla_vc.xlsx: sin datos para {y}-{m:02d}")
        return result

    except Exception as e:
        print(f"  [WARN] planilla_vc.xlsx: {e}")
        import traceback; traceback.print_exc()
        return {}



# ── Template structure detection ─────────────────────────────────────────────
def _find_structure(ws) -> dict:
    """
    Scan a Datos ICP (2) sheet by formula patterns to find:
    - Last data row (col A date)
    - ICP / FIP / Comp VC rows and their next-empty column
    - Monthly return rows and their next-empty column
    Works regardless of which row numbers each template uses.
    """
    # Find last data row
    last_r = 0
    for r in range(ws.max_row, 1, -1):
        v = ws.cell(r, 1).value
        if v and hasattr(v, 'year'):
            last_r = r; break

    jan_row = last_r - 3  # Jan of current year (e.g. 246 if April=249)
    s = {'last_row': last_r, 'next_row': last_r+1, 'jan_row': jan_row}

    # Find VC rows by matching "=+X{jan_row}" in cols P-AB (16-28)
    icp_vc = fip_vc = comp_vc = None
    for r in range(1, 100):
        for c in range(16, 29):
            v = str(ws.cell(r, c).value or '')
            m = re.fullmatch(r'=\+([A-Z]+)' + str(jan_row), v)
            if m:
                src = m.group(1)
                if src == 'B' and icp_vc is None:
                    icp_vc = r; s['icp_vc_row'] = r; s['icp_vc_jan_c'] = c
                elif src in ('G','H','I') and fip_vc is None:
                    fip_vc = r; s['fip_vc_row'] = r; s['fip_vc_jan_c'] = c
                    s['fip_data_col'] = src
                elif src in ('J','K','L') and comp_vc is None:
                    comp_vc = r; s['comp_vc_row'] = r; s['comp_vc_jan_c'] = c
                    s['comp_data_col'] = src

    # Find next-empty column for each VC row
    def next_empty_c(row):
        if row is None: return None, None, None
        for c in range(16, 30):
            if ws.cell(row, c).value is None:
                return c, xl_u.get_column_letter(c), xl_u.get_column_letter(c-1)
        return None, None, None

    s['icp_vc_next_c'], s['icp_vc_next_col'], s['icp_vc_prev_col'] = next_empty_c(icp_vc)
    s['fip_vc_next_c'], s['fip_vc_next_col'], s['fip_vc_prev_col'] = next_empty_c(fip_vc)
    s['comp_vc_next_c'], s['comp_vc_next_col'], s['comp_vc_prev_col'] = next_empty_c(comp_vc)

    # Find monthly return rows by AE column formula matching "P{vc_row}"
    def find_ret_row(vc_row):
        if vc_row is None: return None, None, None, None
        for r in range(1, 100):
            ae = str(ws.cell(r, 31).value or '')  # AE = col 31
            if f'P{vc_row}' in ae:
                for c in range(31, 44):
                    if ws.cell(r, c).value is None:
                        return r, c, xl_u.get_column_letter(c), ws.cell(r, 43).value
                return r, None, None, ws.cell(r, 43).value
        return None, None, None, None

    s['icp_ret_row'],  s['icp_ret_next_c'],  s['icp_ret_next_col'],  s['icp_total']  = find_ret_row(icp_vc)
    s['fip_ret_row'],  s['fip_ret_next_c'],  s['fip_ret_next_col'],  s['fip_total']  = find_ret_row(fip_vc)
    s['comp_ret_row'], s['comp_ret_next_c'], s['comp_ret_next_col'], s['comp_total'] = find_ret_row(comp_vc)

    return s

# ── Update one template ───────────────────────────────────────────────────────
def update_template(tmpl_file: str, y: int, m: int,
                     fip_vc: float, icp_val: float, sant_vc: float) -> str:
    path = _TMPL_DIR / tmpl_file
    if not path.exists():
        return 'skip_not_found'

    wb = openpyxl.load_workbook(path)
    if 'Datos ICP (2)' not in wb.sheetnames:
        wb.close(); return 'skip_no_datos_sheet'

    ws = wb['Datos ICP (2)']
    s  = _find_structure(ws)

    if not s.get('last_row'):
        wb.close(); return 'skip_no_data_rows'

    last_r = s['last_row']
    last_d = ws.cell(last_r, 1).value
    if last_d and hasattr(last_d,'year') and last_d.year == y and last_d.month == m:
        wb.close(); return 'already_updated'

    eom_dt  = datetime(y, m, _eom(y, m).day)
    next_r  = s['next_row']
    next_bB  = next_r  # The new row B reference (ICP)

    # ── 1. Add the new data row (col A-M) ────────────────────────────────────
    ws.cell(next_r, 1).value = eom_dt          # A: date

    # ICP
    ws.cell(next_r, 2).value = icp_val          # B: ICP value
    ws.cell(next_r, 3).value = f"=(B{next_r}/B{last_r}-1)/(A{next_r}-A{last_r})*30"

    # Normalized ICP index
    ws.cell(next_r, 4).value = f"=+D{last_r}*(1+C{next_r})"

    # Fund VC date + raw VC
    ws.cell(next_r, 6).value = eom_dt           # F: date
    ws.cell(next_r, 7).value = fip_vc           # G: raw VC

    # Adjusted fund VC (col H = G + adjustments from fixed rows)
    fip_data_col = s.get('fip_data_col', 'H')
    ws.cell(next_r, 8).value = (
        f"=+G{next_r}+$S$148+$S$150+$S$152+$S$154+$S$156+$S$158"
    )

    # Fund return (col I: normalized to 30-day period)
    ws.cell(next_r, 9).value = (
        f"=+(H{next_r}/H{last_r}-1)/(F{next_r}-F{last_r})*30"
    )

    # Fund normalized index (col J)
    ws.cell(next_r, 10).value = f"=+J{last_r}*(1+I{next_r})"

    # Santander MM VC (col K)
    ws.cell(next_r, 11).value = sant_vc          # K

    # Comp return (col L)
    ws.cell(next_r, 12).value = (
        f"=+(K{next_r}/K{last_r}-1)/(F{next_r}-F{last_r})*30"
    )

    # Comp normalized index (col M)
    ws.cell(next_r, 13).value = f"=+M{last_r}*(1+L{next_r})"

    # ── 2. Update VC lookup rows (add formula for new month) ─────────────────
    icp_vc_row  = s.get('icp_vc_row')
    fip_vc_row  = s.get('fip_vc_row')
    comp_vc_row = s.get('comp_vc_row')
    fip_dc      = s.get('fip_data_col', 'H')
    comp_dc     = s.get('comp_data_col', 'K')

    icp_next_c  = s.get('icp_vc_next_c')
    fip_next_c  = s.get('fip_vc_next_c')
    comp_next_c = s.get('comp_vc_next_c')

    if icp_vc_row and icp_next_c:
        ws.cell(icp_vc_row, icp_next_c).value = f"=+B{next_r}"

    if fip_vc_row and fip_next_c:
        ws.cell(fip_vc_row, fip_next_c).value = f"=+{fip_dc}{next_r}"

    if comp_vc_row and comp_next_c:
        ws.cell(comp_vc_row, comp_next_c).value = f"=+{comp_dc}{next_r}"

    # ── 3. Add monthly return formulas ────────────────────────────────────────
    icp_ret_row  = s.get('icp_ret_row')
    fip_ret_row  = s.get('fip_ret_row')
    comp_ret_row = s.get('comp_ret_row')

    icp_ret_c    = s.get('icp_ret_next_c')
    fip_ret_c    = s.get('fip_ret_next_c')
    comp_ret_c   = s.get('comp_ret_next_c')

    icp_next_col  = s.get('icp_vc_next_col', 'T')
    icp_prev_col  = s.get('icp_vc_prev_col', 'S')
    fip_next_col  = s.get('fip_vc_next_col', 'T')
    fip_prev_col  = s.get('fip_vc_prev_col', 'S')
    comp_next_col = s.get('comp_vc_next_col', 'T')
    comp_prev_col = s.get('comp_vc_prev_col', 'S')

    if icp_ret_row and icp_ret_c and icp_vc_row:
        ws.cell(icp_ret_row, icp_ret_c).value = (
            f"=+{icp_next_col}{icp_vc_row}/{icp_prev_col}{icp_vc_row}-1"
        )

    if fip_ret_row and fip_ret_c and fip_vc_row:
        ws.cell(fip_ret_row, fip_ret_c).value = (
            f"=+{fip_next_col}{fip_vc_row}/{fip_prev_col}{fip_vc_row}-1"
        )

    if comp_ret_row and comp_ret_c and comp_vc_row:
        ws.cell(comp_ret_row, comp_ret_c).value = (
            f"=+{comp_next_col}{comp_vc_row}/{comp_prev_col}{comp_vc_row}-1"
        )

    # ── 4. Update total (AQ col) formulas to include new month ───────────────
    # Pattern: =+(S23/P23-1)/COUNTA(AE22:AP22)*12
    # Update the last VC reference (S→T) and the COUNTA range (AE22:AP22 → AE22:AQ22)
    def update_total(ret_row, vc_row, new_vc_col, new_ret_col):
        if not (ret_row and vc_row and new_vc_col and new_ret_col): return
        old_total = ws.cell(ret_row, 43).value
        if not old_total or not isinstance(old_total, str): return
        # Replace the leading VC reference (e.g. S23 → T23)
        new_total = re.sub(
            r'[A-Z]+' + str(vc_row) + r'/',
            f"{new_vc_col}{vc_row}/",
            old_total, count=1
        )
        # Replace COUNTA end reference (e.g. AP22 → AQ22)
        new_total = re.sub(
            r'COUNTA\(AE{r}:([A-Z]+){r}\)'.format(r=ret_row),
            f"COUNTA(AE{ret_row}:{new_ret_col}{ret_row})",
            new_total
        )
        ws.cell(ret_row, 43).value = new_total

    icp_ret_col  = s.get('icp_ret_next_col', 'AI')
    fip_ret_col  = s.get('fip_ret_next_col', 'AI')
    comp_ret_col = s.get('comp_ret_next_col', 'AI')

    update_total(icp_ret_row,  icp_vc_row,  icp_next_col,  icp_ret_col)
    update_total(fip_ret_row,  fip_vc_row,  fip_next_col,  fip_ret_col)
    update_total(comp_ret_row, comp_vc_row, comp_next_col, comp_ret_col)

    # ── 5. Update AW19 (current month date cell) ─────────────────────────────
    ws.cell(19, 49).value = eom_dt  # AW19

    # ── 6. Update rentabilidad sheet Acum header ──────────────────────────────
    if 'rentabilidad' in wb.sheetnames:
        wb['rentabilidad'].cell(1, 24).value = f"Acum\n{y} (*)"

    wb.save(path)
    wb.close()
    return 'updated'


# ── Main entry point ──────────────────────────────────────────────────────────
def run_update(target_year: int = None, target_month: int = None) -> dict:
    if not target_year:
        tm = os.environ.get("TARGET_MONTH", "").strip()
        if tm:
            target_year, target_month = map(int, tm.split("-"))
        else:
            hoy  = date.today()
            prev = date(hoy.year, hoy.month, 1) - relativedelta(months=1)
            target_year, target_month = prev.year, prev.month

    y, m = target_year, target_month
    print(f"\n{'='*60}")
    print(f" Actualizando templates para {y}-{m:02d}")
    print(f"{'='*60}\n")

    icp_val  = _get_icp(y, m)
    sant_vc  = _get_santander(y, m)
    planilla = _read_planilla(y, m)

    print(f"  ICP {y}-{m:02d}:       {icp_val}")
    print(f"  Santander MM:  {sant_vc}")
    print(f"  Planilla VC:   {len(planilla)} fondos\n")

    results = {}
    for tmpl_file, ods_name in FUND_ODS_NAMES.items():
        # Get fund VC
        fip_vc = None

        # 1. From planilla - try exact match first, then normalized
        if ods_name in planilla:
            fip_vc = planilla[ods_name]
        elif ods_name.replace('Ó','O').replace('ó','o') in {k.replace('Ó','O').replace('ó','o'): v for k,v in planilla.items()}:
            norm = {k.replace('Ó','O').replace('ó','o'): v for k,v in planilla.items()}
            fip_vc = norm.get(ods_name.replace('Ó','O').replace('ó','o'))
        elif ods_name.upper() in {k.upper(): v for k,v in planilla.items()}:
            fip_vc = {k.upper(): v for k,v in planilla.items()}.get(ods_name.upper())

        # 2. From ODS
        if fip_vc is None:
            fip_vc = _ods_vc(ods_name, y, m)

        if fip_vc is None:
            print(f"  [SKIP] {tmpl_file}: no VC disponible para {y}-{m:02d}")
            results[tmpl_file] = 'skip_no_data'
            continue

        if icp_val is None or sant_vc is None:
            print(f"  [SKIP] {tmpl_file}: falta ICP o Santander MM")
            results[tmpl_file] = 'skip_missing_deps'
            continue

        result = update_template(tmpl_file, y, m, fip_vc, icp_val, sant_vc)
        status_str = {
            'updated':       f"✓ {tmpl_file} — fip_vc={fip_vc:.4f}",
            'already_updated': f"— {tmpl_file}: ya actualizado",
            'skip_not_found':  f"✗ {tmpl_file}: archivo no encontrado",
        }.get(result, f"  {result}: {tmpl_file}")
        print(f"  {status_str}")
        results[tmpl_file] = result

    updated = sum(1 for v in results.values() if v == 'updated')
    skipped = sum(1 for v in results.values() if 'skip' in v or v == 'already_updated')
    print(f"\n  Resumen: {updated} actualizados, {skipped} omitidos\n")
    return results


if __name__ == "__main__":
    run_update()
