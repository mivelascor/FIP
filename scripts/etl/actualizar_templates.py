"""
etl/actualizar_templates.py

Para cada fondo, actualiza el template Excel con los datos del mes nuevo:
  - Inserta nueva fila en 'Datos ICP (2)': fecha, ICP, VC fondo, VC comp + fórmulas
  - Actualiza celda AW19/AW20 con la nueva fecha
  - Ejecuta recalc.py para que LibreOffice recalcule todas las fórmulas
  → La hoja 'rentabilidad' queda con los valores exactos listos para template_reader.py

Fuentes de datos:
  - ICP: mindicador.cl (o BCCh con credenciales)
  - VC fondo: API SQL ODS.VALORES_CUOTA_GPI
  - VC comp CLP: CMF (Santander Money Market)
  - VC comp USD: CMF (BanChile Corporate Dollar)
"""
import re
import sys
import subprocess
import requests
import pandas as pd
from datetime import date, timedelta, datetime
from pathlib import Path
from openpyxl import load_workbook

# Ruta al script recalc.py (se copia del skill al setup)
RECALC_SCRIPT = Path(__file__).parent.parent / "scripts_lib" / "recalc.py"
TEMPLATES_DIR  = Path(__file__).parent.parent.parent / "inputs" / "templates"

MINDICADOR = "https://mindicador.cl/api/tpm"
API_SQL    = "https://claudeods.vantrustcapital.cl/query"

CMF_URL_CLP = ("https://www.cmfchile.cl/institucional/mercados/entidad.php"
               "?mercado=V&rut=8057&grupo=&tipoentidad=RGFMU"
               "&row=AAAw%20cAAhAAAACcAAs&vig=VI&control=svs&pestania=7")
CMF_URL_USD = ("https://www.cmfchile.cl/institucional/mercados/entidad.php"
               "?mercado=V&rut=8248&grupo=&tipoentidad=RGFMU"
               "&row=AAAw%20cAAhAAAACfAAj&vig=VI&control=svs&pestania=7")

TEMPLATE_MAP = {
    "FIP VANTRUST LIQUIDEZ ACTIVA":        "TEMPLATE_FONDO_LIQUIDEZ_ACTIVA.xlsx",
    "FIP VANTRUST LIQUIDEZ ALTO APORTE":   "TEMPLATE_FONDO_ALTO_APORTE.xlsx",
    "FIP VANTRUST LIQUIDEZ ALTO CAPITAL":  "TEMPLATE_FONDO_ALTO_CAPITAL.xlsx",
    "FIP VANTRUST LIQUIDEZ ALTO MONTO":    "TEMPLATE_FONDO_LIQUIDEZ_ALTO_MONTO.xlsx",
    "FIP VANTRUST LIQUIDEZ CAJA":          "TEMPLATE_FONDO_LIQUIDEZ_CAJA.xlsx",
    "FIP VANTRUST LIQUIDEZ CONTINUA":      "TEMPLATE_FONDO_LIQUIDEZ_CONTINUA.xlsx",
    "FIP VANTRUST LIQUIDEZ CORRIENTE":     "TEMPLATE_FONDO_LIQUIDEZ_CORRIENTE.xlsx",
    "FIP VANTRUST LIQUIDEZ CORTO PLAZO":   "TEMPLATE_FONDO_LIQUIDEZ_CORTO_PLAZO.xlsx",
    "FIP VANTRUST LIQUIDEZ DISPONIBLE I":  "TEMPLATE_FONDO_LIQUIDEZ_Disponible_I.xlsx",
    "FIP VANTRUST LIQUIDEZ DOLAR":         "TEMPLATE_FONDO_LIQUIDEZ_DOLAR.xlsx",
    "FIP VANTRUST LIQUIDEZ DOLAR CAJA":    "TEMPLATE_FONDO_LIQUIDEZ_DOLAR_CAJA.xlsx",
    "FIP VANTRUST LIQUIDEZ EFECTIVO":      "TEMPLATE_FONDO_LIQUIDEZ_EFECTIVO.xlsx",
    "FIP VANTRUST LIQUIDEZ FLEXIBLE":      "TEMPLATE_FONDO_LIQUIDEZ_FLEXIBLE.xlsx",
    "FIP VANTRUST LIQUIDEZ I":             "TEMPLATE_FONDO_LIQUIDEZ_UNO.xlsx",
    "FIP VANTRUST LIQUIDEZ LOCAL":         "TEMPLATE_FONDO_LIQUIDEZ_LOCAL.xlsx",
    "FIP VANTRUST LIQUIDEZ MONETARIO I":   "TEMPLATE_FONDO_LIQUIDEZ_Monetario_I.xlsx",
    "FIP VANTRUST LIQUIDEZ PERMANENTE":    "TEMPLATE_FONDO_LIQUIDEZ_Permanente.xlsx",
    "FIP VANTRUST LIQUIDEZ PLUS":          "TEMPLATE_FONDO_LIQUIDEZ_PLUS.xlsx",
    "FIP VANTRUST LIQUIDEZ PRESENTE":      "TEMPLATE_FONDO_LIQUIDEZ_Presente.xlsx",
    "FIP VANTRUST LIQUIDEZ RECURRENTE":    "TEMPLATE_FONDO_LIQUIDEZ.xlsx",
    "FIP VANTRUST LIQUIDEZ RENDIMIENTO":   "TEMPLATE_FONDO_LIQUIDEZ_RENDIMIENTO.xlsx",
    "FIP VANTRUST LIQUIDEZ RESERVA DOLAR": "TEMPLATE_FONDO_LIQUIDEZ_RESERVA_DOLAR.xlsx",
    "FIP VANTRUST LIQUIDEZ SENCILLO":      "TEMPLATE_FONDO_LIQUIDEZ_SENCILLO.xlsx",
    "FIP VANTRUST LIQUIDEZ TEMPORAL":      "TEMPLATE_FONDO_USD_MONEY_MARKET.xlsx",
}

# LIQUIDEZ I usa AW20 en vez de AW19
CELDA_FECHA_ROW = {"FIP VANTRUST LIQUIDEZ I": 20}


# ── Obtener datos ───────────────────────────────────────────────────────────

def get_icp_nivel(fecha_obj: datetime) -> float | None:
    """Nivel acumulado del ICP para el mes de fecha_obj."""
    all_data = []
    for y in range(2009, fecha_obj.year + 1):
        try:
            r = requests.get(f"{MINDICADOR}/{y}", timeout=15)
            for item in r.json().get("serie", []):
                all_data.append({"fecha": pd.to_datetime(item["fecha"]),
                                  "tpm": float(item["valor"])})
        except Exception:
            continue
    if not all_data:
        return None
    df = pd.DataFrame(all_data).sort_values("fecha").reset_index(drop=True)
    df["period"] = df["fecha"].dt.to_period("M")
    m  = df.groupby("period")["tpm"].mean().reset_index()
    m["rent"]  = m["tpm"] / 1200
    nivel = [10000.0]
    for i in range(1, len(m)):
        nivel.append(nivel[-1] * (1 + m.loc[i, "rent"]))
    m["nivel"] = nivel
    m["fecha"]  = m["period"].dt.to_timestamp("M")
    s = pd.Series(m["nivel"].values, index=pd.DatetimeIndex(m["fecha"])).sort_index()
    target = pd.Timestamp(fecha_obj.year, fecha_obj.month, 1) + pd.offsets.MonthEnd(0)
    cerca  = s[s.index <= target]
    return float(cerca.iloc[-1]) if not cerca.empty else None


def get_vc_fondo(nombre_fondo: str, fecha_obj: datetime) -> float | None:
    """Valor cuota del fondo para fecha_obj desde la API SQL."""
    desde = (fecha_obj - timedelta(days=45)).strftime("%Y-%m-%d")
    hasta = fecha_obj.strftime("%Y-%m-%d")
    sql = (f"SELECT FECHA_CIERRE AS fecha, VALOR_CUOTA "
           f"FROM ODS.VALORES_CUOTA_GPI "
           f"WHERE FECHA_CIERRE BETWEEN '{desde}' AND '{hasta}' "
           f"AND RTRIM(LTRIM(EMPRESA)) = '{nombre_fondo}' "
           f"AND VALOR_CUOTA > 0 ORDER BY FECHA_CIERRE DESC")
    try:
        r = requests.post(API_SQL, json={"Sql": sql},
                          headers={"Content-Type":"application/json"}, timeout=60)
        r.raise_for_status()
        data = r.json()
        rows = data if isinstance(data, list) else data.get("rows", data.get("data", []))
        if not rows:
            return None
        df = pd.DataFrame(rows)
        df.columns = [c.lower() for c in df.columns]
        for col in ("fecha_cierre","fecha"):
            if col in df.columns: df = df.rename(columns={col:"fecha"}); break
        for col in ("valor_cuota","precio","valor"):
            if col in df.columns: df = df.rename(columns={col:"valor_cuota"}); break
        df["fecha"]       = pd.to_datetime(df["fecha"])
        df["valor_cuota"] = pd.to_numeric(df["valor_cuota"], errors="coerce")
        df = df.dropna().sort_values("fecha")
        target = pd.Timestamp(fecha_obj)
        if target in df["fecha"].values:
            return float(df[df["fecha"] == target]["valor_cuota"].iloc[0])
        return float(df.iloc[-1]["valor_cuota"])
    except Exception as e:
        print(f"    [WARN] API SQL ({nombre_fondo}): {e}")
        return None


def scrape_cmf(url: str) -> tuple:
    """Último valor cuota y fecha de la CMF."""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            b = p.chromium.launch(headless=True)
            page = b.new_page()
            page.goto(url, timeout=30000)
            page.wait_for_timeout(3000)
            val, fecha = None, None
            for row in page.query_selector_all("table tr"):
                cells = row.query_selector_all("td")
                if len(cells) >= 2:
                    try:
                        f = cells[0].inner_text().strip()
                        v = float(cells[1].inner_text().strip().replace(",","."))
                        if v > 0: val, fecha = v, f
                    except Exception: continue
            b.close()
            return val, fecha
    except Exception as e:
        print(f"    [WARN] CMF scraping: {e}")
        return None, None


# ── Actualizar Excel ────────────────────────────────────────────────────────

def _inc_formula(formula: str, old_row: int, new_row: int) -> str:
    """Incrementa referencias de fila relativas en una fórmula."""
    if not formula or not formula.startswith("="):
        return formula
    def replace(m):
        col, row = m.group(1), int(m.group(2))
        if "$" in col: return m.group(0)  # referencia absoluta de fila: no cambiar
        if row == old_row:     return f"{col}{new_row}"
        if row == old_row - 1: return f"{col}{new_row - 1}"
        return m.group(0)
    return re.sub(r'([A-Za-z]+\$?)(\d+)', replace, formula)


def _recalcular(ruta: Path):
    """Ejecuta recalc.py para que LibreOffice recalcule las fórmulas del Excel."""
    if not RECALC_SCRIPT.exists():
        # Intentar con libreoffice directo
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "xlsx",
             "--outdir", str(ruta.parent), str(ruta)],
            capture_output=True, timeout=120
        )
        return
    result = subprocess.run(
        [sys.executable, str(RECALC_SCRIPT), str(ruta), "120"],
        capture_output=True, text=True, timeout=150,
        cwd=str(RECALC_SCRIPT.parent)
    )
    if result.returncode != 0:
        print(f"    [WARN] recalc error: {result.stderr[:200]}")


def actualizar_template(nombre_fondo: str, fecha_obj: datetime,
                         icp_nivel: float, vc_fondo: float,
                         vc_comp: float) -> bool:
    """Actualiza el template del fondo con los datos del mes nuevo."""
    archivo = TEMPLATE_MAP.get(nombre_fondo)
    if not archivo:
        return False
    ruta = TEMPLATES_DIR / archivo
    if not ruta.exists():
        print(f"    [WARN] No encontrado: {ruta}")
        return False

    wb = load_workbook(str(ruta), data_only=False)
    ws = wb["Datos ICP (2)"]

    # Encontrar la última fila con datos en col A
    last_row = None
    for row in range(1, 500):
        v = ws.cell(row, 1).value
        if v and hasattr(v, "year"):
            last_row = row
    if last_row is None:
        wb.close(); return False

    # Verificar si ya tiene este mes
    last_date = ws.cell(last_row, 1).value
    if (last_date and hasattr(last_date, "year") and
            last_date.year == fecha_obj.year and last_date.month == fecha_obj.month):
        print(f"    Ya actualizado ({fecha_obj.strftime('%Y-%m')})")
        wb.close(); return True

    new_row = last_row + 1

    # Insertar datos del mes nuevo
    ws.cell(new_row, 1).value  = fecha_obj   # col A: fecha
    ws.cell(new_row, 2).value  = icp_nivel   # col B: ICP
    ws.cell(new_row, 6).value  = fecha_obj   # col F: fecha (igual a A)
    ws.cell(new_row, 7).value  = vc_fondo    # col G: VC fondo
    if vc_comp is not None:
        ws.cell(new_row, 11).value = vc_comp # col K: VC competencia

    # Copiar e incrementar fórmulas C,D,H,I,J,L,M de la fila anterior
    for col in (3, 4, 8, 9, 10, 12, 13):
        formula_prev = ws.cell(last_row, col).value
        if formula_prev and isinstance(formula_prev, str) and formula_prev.startswith("="):
            ws.cell(new_row, col).value = _inc_formula(formula_prev, last_row, new_row)

    # Actualizar celda de fecha AW19/AW20
    fecha_row = CELDA_FECHA_ROW.get(nombre_fondo, 19)
    ws.cell(fecha_row, 49).value = fecha_obj  # col AW (49)

    wb.save(str(ruta))
    wb.close()

    # Recalcular con LibreOffice
    _recalcular(ruta)
    return True


def actualizar_todos(fecha_obj: datetime, fondos: list) -> dict:
    """
    Actualiza todos los templates en paralelo (ICP y competencia se obtienen una vez).
    Retorna dict {nombre_fondo: True/False}.
    """
    print("  Obteniendo ICP...")
    icp_nivel = get_icp_nivel(fecha_obj)
    if icp_nivel is None:
        print("  [ERROR] No se pudo obtener el ICP")
        return {f: False for f in fondos}
    print(f"    ICP {fecha_obj.strftime('%Y-%m')}: {icp_nivel:.2f}")

    print("  Obteniendo competencia CLP (CMF)...")
    vc_comp_clp, _ = scrape_cmf(CMF_URL_CLP)
    print(f"    CLP: {vc_comp_clp}")

    print("  Obteniendo competencia USD (CMF)...")
    vc_comp_usd, _ = scrape_cmf(CMF_URL_USD)
    print(f"    USD: {vc_comp_usd}")

    resultados = {}
    for nombre_fondo in fondos:
        print(f"\n  → {nombre_fondo}")
        es_usd   = any(x in nombre_fondo.upper() for x in ("DOLAR", "USD"))
        vc_comp  = vc_comp_usd if es_usd else vc_comp_clp

        vc_fondo = get_vc_fondo(nombre_fondo, fecha_obj)
        if vc_fondo is None:
            print(f"    [WARN] Sin VC del fondo")
            resultados[nombre_fondo] = False
            continue

        print(f"    VC fondo: {vc_fondo:.4f}")
        try:
            ok = actualizar_template(nombre_fondo, fecha_obj,
                                     icp_nivel, vc_fondo, vc_comp)
            resultados[nombre_fondo] = ok
        except Exception as e:
            print(f"    [ERROR]: {e}")
            resultados[nombre_fondo] = False

    return resultados
