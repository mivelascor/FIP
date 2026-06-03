"""
etl/template_reader.py  —  Fuente definitiva de datos para folletos Vantrust.

REGLAS (verificadas contra PDFs manuales):
─ CLP: resumen M/T/S/A/Ac calculado desde ODS VC (simple return).
  - Para meses antes de ODS (jul-2025), se encadena VC desde template historico.
─ USD: resumen y histórico leídos directamente del template rentabilidad sheet
  (el ODS tiene ligeras diferencias por fuente; el template es más exacto).
─ ICP: serie CLICP desde icp_clicp.json + extensión TPM para meses futuros.
─ Competencia CLP: Santander MM desde comp_clp.json.
─ Competencia USD: leída del template (col K = Banchile Corporate Dollar normalizado).
"""
import os, requests, json, calendar
from pathlib import Path
from datetime import date
from dateutil.relativedelta import relativedelta
import pandas as pd

_INPUTS_DIR = Path(__file__).parent.parent.parent / "inputs"
_TMPL_DIR   = _INPUTS_DIR / "templates"

# ── Mapa: ODS nombre → template file con datos históricos ────────────────────
FUND_TEMPLATE_MAP = {
    # Direct 1:1 mapping — each template has data for its own fund
    'FIP VANTRUST LIQUIDEZ ACTIVA':        'TEMPLATE FONDO LIQUIDEZ ACTIVA.xlsx',
    'FIP VANTRUST LIQUIDEZ ALTO APORTE':   'TEMPLATE FONDO ALTO APORTE.xlsx',
    'FIP VANTRUST LIQUIDEZ ALTO CAPITAL':  'TEMPLATE FONDO ALTO CAPITAL.xlsx',
    'FIP VANTRUST LIQUIDEZ ALTO MONTO':    'TEMPLATE FONDO LIQUIDEZ ALTO MONTO.xlsx',
    'FIP VANTRUST LIQUIDEZ CAJA':          'TEMPLATE FONDO LIQUIDEZ CAJA.xlsx',
    'FIP VANTRUST LIQUIDEZ CONTINUA':      'TEMPLATE FONDO LIQUIDEZ CONTINUA.xlsx',
    'FIP VANTRUST LIQUIDEZ CORRIENTE':     'TEMPLATE FONDO LIQUIDEZ CORRIENTE.xlsx',
    'FIP VANTRUST LIQUIDEZ CORTO PLAZO':   'TEMPLATE FONDO LIQUIDEZ CORTO PLAZO.xlsx',
    'FIP VANTRUST LIQUIDEZ DISPONIBLE I':  'TEMPLATE FONDO LIQUIDEZ Disponible I.xlsx',
    'FIP VANTRUST LIQUIDEZ DOLAR':         'TEMPLATE FONDO LIQUIDEZ DOLAR.xlsx',
    'FIP VANTRUST LIQUIDEZ DOLAR CAJA':    'TEMPLATE FONDO LIQUIDEZ DOLAR CAJA.xlsx',
    'FIP VANTRUST LIQUIDEZ EFECTIVO':      'TEMPLATE FONDO LIQUIDEZ EFECTIVO.xlsx',
    'FIP VANTRUST LIQUIDEZ FLEXIBLE':      'TEMPLATE FONDO LIQUIDEZ FLEXIBLE.xlsx',
    'FIP VANTRUST LIQUIDEZ I':             'TEMPLATE FONDO LIQUIDEZ UNO.xlsx',
    'FIP VANTRUST LIQUIDEZ LOCAL':         'TEMPLATE FONDO LIQUIDEZ LOCAL.xlsx',
    'FIP VANTRUST LIQUIDEZ MONETARIO I':   'TEMPLATE FONDO LIQUIDEZ Monetario I.xlsx',
    'FIP VANTRUST LIQUIDEZ PERMANENTE':    'TEMPLATE FONDO LIQUIDEZ Permanente.xlsx',
    'FIP VANTRUST LIQUIDEZ PLUS':          'TEMPLATE FONDO LIQUIDEZ PLUS.xlsx',
    'FIP VANTRUST LIQUIDEZ PRESENTE':      'TEMPLATE FONDO LIQUIDEZ Presente.xlsx',
    'FIP VANTRUST LIQUIDEZ RENDIMIENTO':   'TEMPLATE FONDO LIQUIDEZ RENDIMIENTO.xlsx',
    'FIP VANTRUST LIQUIDEZ RESERVA DÓLAR': 'TEMPLATE FONDO LIQUIDEZ RESERVA DOLAR.xlsx',
    'FIP VANTRUST LIQUIDEZ SENCILLO':      'TEMPLATE FONDO LIQUIDEZ SENCILLO.xlsx',
}

FONDOS_USD = {
    'FIP VANTRUST LIQUIDEZ DOLAR',
    'FIP VANTRUST LIQUIDEZ DOLAR CAJA',
    'FIP VANTRUST LIQUIDEZ RESERVA DÓLAR',
}

# Funds where template summary is used (instead of ODS calculation).
# This includes ALL funds that have FIP data in their template rentabilidad sheet.
# USD funds: always use template (different formula + data source)
# CLP funds with template: use template (SharePoint data matches PDFs better than ODS)
# Only USD funds use template for summary.
# CLP funds always use ODS for summary (templates may be stale by 1-2 months).
# Templates are used for historico (pre-ODS months) for ALL funds in FUND_TEMPLATE_MAP.
FONDOS_USE_TEMPLATE = FONDOS_USD

NOMBRE_DISPLAY = {
    'FIP VANTRUST LIQUIDEZ ACTIVA':        'FIP Liquidez Activa',
    'FIP VANTRUST LIQUIDEZ ALTO APORTE':   'FIP Liquidez Alto Aporte',
    'FIP VANTRUST LIQUIDEZ ALTO CAPITAL':  'FIP Alto Capital',
    'FIP VANTRUST LIQUIDEZ ALTO MONTO':    'FIP Liquidez Alto Monto',
    'FIP VANTRUST LIQUIDEZ CAJA':          'FIP Liquidez Caja',
    'FIP VANTRUST LIQUIDEZ CONTINUA':      'FIP Liquidez Continua',
    'FIP VANTRUST LIQUIDEZ CORRIENTE':     'FIP Liquidez Corriente',
    'FIP VANTRUST LIQUIDEZ CORTO PLAZO':   'FIP Liquidez Corto Plazo',
    'FIP VANTRUST LIQUIDEZ DISPONIBLE I':  'FIP Liquidez Disponible I',
    'FIP VANTRUST LIQUIDEZ DOLAR':         'FIP Liquidez Dólar',
    'FIP VANTRUST LIQUIDEZ DOLAR CAJA':    'FIP Liquidez Dólar Caja',
    'FIP VANTRUST LIQUIDEZ EFECTIVO':      'FIP Liquidez Efectivo',
    'FIP VANTRUST LIQUIDEZ FLEXIBLE':      'FIP Liquidez Flexible',
    'FIP VANTRUST LIQUIDEZ I':             'FIP Liquidez I',
    'FIP VANTRUST LIQUIDEZ LOCAL':         'FIP Liquidez Local',
    'FIP VANTRUST LIQUIDEZ MONETARIO I':   'FIP Liquidez Monetario I',
    'FIP VANTRUST LIQUIDEZ PERMANENTE':    'FIP Liquidez Permanente',
    'FIP VANTRUST LIQUIDEZ PLUS':          'FIP Liquidez Plus',
    'FIP VANTRUST LIQUIDEZ PRESENTE':      'FIP Liquidez Presente',
    'FIP VANTRUST LIQUIDEZ RECURRENTE':    'FIP Liquidez Recurrente',
    'FIP VANTRUST LIQUIDEZ RENDIMIENTO':   'FIP Liquidez Rendimiento',
    'FIP VANTRUST LIQUIDEZ RESERVA DÓLAR': 'FIP Liquidez Reserva Dólar',
    'FIP VANTRUST LIQUIDEZ SENCILLO':      'FIP Liquidez Sencillo',
    'FIP VANTRUST LIQUIDEZ TEMPORAL':      'FIP Liquidez Temporal',
}

_ICP_CACHE:  dict = {}
_VC_CACHE:   dict = {}
_HIST_CACHE: dict = {}
_ODS_API = "https://claudeods.vantrustcapital.cl/query"

# ── Helpers ───────────────────────────────────────────────────────────────────
def _load_json(fn):
    try:
        with open(_INPUTS_DIR / fn) as f:
            return {(int(k[:4]), int(k[5:7])): float(v) for k,v in json.load(f).items()}
    except Exception as e:
        print(f"  [WARN] {fn}: {e}"); return {}

def _eom(y, m):
    return date(y, m, calendar.monthrange(y, m)[1])

def _prev(y, m, n=1):
    d = date(y, m, 1) - relativedelta(months=n)
    return d.year, d.month

# ── ICP ───────────────────────────────────────────────────────────────────────
def _get_icp_series():
    global _ICP_CACHE
    if _ICP_CACHE: return _ICP_CACHE
    clicp = _load_json("icp_clicp.json")
    after, base = (max(clicp.keys()), clicp[max(clicp.keys())]) if clicp else ((2005,12), 10000.0)
    if not _extend_bcch(clicp, after, base):
        _extend_tpm(clicp, after, base)
    _ICP_CACHE = clicp
    return _ICP_CACHE

def _extend_tpm(series, after_ym, base_val):
    prev = base_val
    for year in range(after_ym[0], date.today().year + 1):
        try:
            r  = requests.get(f"https://mindicador.cl/api/tpm/{year}", timeout=15)
            df = pd.DataFrame([{"fecha": pd.to_datetime(i["fecha"]), "tpm": float(i["valor"])}
                                for i in r.json().get("serie", [])])
            if df.empty: continue
            df["ym"] = df["fecha"].dt.to_period("M")
            for _, row in df.groupby("ym")["tpm"].mean().reset_index().iterrows():
                ym = (row["ym"].year, row["ym"].month)
                if ym <= after_ym: continue
                prev = prev * (1 + row["tpm"] / 1200)
                series[ym] = prev
        except: pass

def _extend_bcch(series, after_ym, base_val):
    """Extiende la serie ICP desde after_ym usando la TIB diaria del BCCh,
    compuesta sobre TODOS los dias calendario (forward-fill de la tasa).
    Ancla en el ultimo valor oficial del JSON; solo agrega meses COMPLETOS.
    Devuelve True si extendio (o no habia nada que extender) usando BCCh; False si no hay credenciales/datos."""
    user = os.environ.get("BCCH_USER",""); pwd = os.environ.get("BCCH_PASS","")
    if not (user and pwd): return False
    import calendar as _cal
    from datetime import timedelta as _td
    try:
        start = f"{after_ym[0]}-{after_ym[1]:02d}-01"
        params = {"user":user,"pass":pwd,"function":"GetSeries",
                  "timeseries":"F022.TIB.TIP.D001.NO.Z.D",
                  "firstdate":start,"lastdate":date.today().strftime("%Y-%m-%d")}
        d = requests.get("https://si3.bcentral.cl/SieteRestWS/SieteRestWS.ashx",
                         params=params, timeout=30).json()
        if d.get("Codigo")!=0: return False
        rate={}
        for o in d["Series"]["Obs"]:
            v=o.get("value","")
            if v and v not in ("","NaN","ND"):
                rate[pd.to_datetime(o["indexDateString"],dayfirst=True).date()]=float(v)
        if not rate: return False
        keys=sorted(rate); last_day=keys[-1]
        def rate_on(dd):
            prev=[k for k in keys if k<=dd]
            return rate[prev[-1]] if prev else None
        prev_val=base_val; cy,cm = (after_ym[0]+1,1) if after_ym[1]==12 else (after_ym[0],after_ym[1]+1)
        today=date.today()
        while (cy,cm) <= (today.year, today.month):
            eom=date(cy,cm,_cal.monthrange(cy,cm)[1])
            if last_day < eom: break   # mes incompleto: no extender
            nivel=prev_val; dd=date(cy,cm,1)
            while dd<=eom:
                r=rate_on(dd)
                if r is not None: nivel*=(1+r/100/360)
                dd+=_td(days=1)
            series[(cy,cm)]=nivel; prev_val=nivel
            print(f"    [ICP] {cy}-{cm:02d} desde BCCh: {nivel:.2f}")
            cy,cm = (cy+1,1) if cm==12 else (cy,cm+1)
        return True
    except Exception as e:
        print(f"    [WARN] BCCh ICP no disponible: {e}")
        return False

def _comp_clp_con_cmf():
    """Carga comp_clp.json y, si falta el mes objetivo, lo agrega scrapeando CMF
    (Santander MM serie UNIVE). Si CMF falla, usa el fallback del scraper (extrapolacion)."""
    vc = _load_json("comp_clp.json")
    try:
        from etl.cmf_scraper import get_competencia_clp
        df, val_nuevo, fecha_nueva = get_competencia_clp()
        for _, row in df.iterrows():
            ts = pd.to_datetime(row["fecha"]); key=(ts.year, ts.month)
            if key not in vc:
                vc[key]=float(row["valor_cuota"])
                print(f"    [CMF] comp CLP {key} = {vc[key]:.4f}")
    except Exception as e:
        print(f"    [WARN] CMF comp no disponible: {e}")
    return vc

# ── ODS fetch ─────────────────────────────────────────────────────────────────
def _get_ods_vc(nombre):
    if nombre in _VC_CACHE: return _VC_CACHE[nombre]
    for _ in range(3):
        try:
            sql = (f"SELECT YEAR(FECHA_CIERRE) yr, MONTH(FECHA_CIERRE) mo, MAX(VALOR_CUOTA) vc "
                   f"FROM ODS.VALORES_CUOTA_GPI "
                   f"WHERE RTRIM(LTRIM(EMPRESA))=N'{nombre}' AND VALOR_CUOTA>0 "
                   f"GROUP BY YEAR(FECHA_CIERRE), MONTH(FECHA_CIERRE) ORDER BY yr, mo")
            rows = requests.post(_ODS_API, json={"Sql":sql},
                                 headers={"Content-Type":"application/json"}, timeout=30
                                 ).json().get("rows", [])
            result = {(int(x["yr"]),int(x["mo"])): float(x["vc"]) for x in rows}
            _VC_CACHE[nombre] = result; return result
        except Exception as e:
            import time; time.sleep(1)
    print(f"  [WARN] ODS failed: {nombre}"); return {}

# ── Template historico reader ─────────────────────────────────────────────────
def _get_tmpl_hist(tmpl_file):
    """
    Load historico data from template by reading RAW DATA from Datos ICP sheet.
    Reads cols A(date), B(ICP), G(FIP raw VC), K(Santander VC) directly —
    bypasses formula cells (which have None cached values) entirely.
    Returns {year: {label: {'months': [Jan..Dec returns], 'total': annualized}}}
    """
    if tmpl_file in _HIST_CACHE: return _HIST_CACHE[tmpl_file]
    import openpyxl
    path = _TMPL_DIR / tmpl_file
    if not path.exists(): return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Datos ICP (2)' not in wb.sheetnames:
        wb.close(); return {}

    ws = wb['Datos ICP (2)']

    # ── Step 1: Read all raw VC series from data rows ────────────────────────
    icp_vc  = {}  # {(yr,mo): ICP level}
    fip_vc  = {}  # {(yr,mo): FIP VC (G col = raw)}
    comp_vc = {}  # {(yr,mo): Santander VC (K col)}

    # Detect which column has the FIP fund VC
    # Try G first; if mostly None, try I (some templates use col I for adjusted)
    for r in range(3, ws.max_row + 1):
        date_v = ws.cell(r, 1).value
        if not (date_v and hasattr(date_v, 'year')): continue
        ym = (date_v.year, date_v.month)
        b = ws.cell(r, 2).value   # ICP
        g = ws.cell(r, 7).value   # FIP raw VC (col G)
        i_ = ws.cell(r, 9).value  # FIP adjusted VC (col I, some templates)
        k = ws.cell(r, 11).value  # Santander MM

        if b and isinstance(b, (int, float)):   icp_vc[ym]  = float(b)
        if g and isinstance(g, (int, float)):   fip_vc[ym]  = float(g)
        elif i_ and isinstance(i_, (int, float)): fip_vc[ym] = float(i_)
        if k and isinstance(k, (int, float)):   comp_vc[ym] = float(k)

    wb.close()

    if not fip_vc:
        _HIST_CACHE[tmpl_file] = {}
        return {}

    # ── Step 2: Compute monthly simple returns from VC series ─────────────────
    def monthly_rets(vc_dict):
        out = {}
        for (yr, mo), vc_end in sorted(vc_dict.items()):
            prev_ym = (yr-1, 12) if mo == 1 else (yr, mo-1)
            vc_prev = vc_dict.get(prev_ym)
            out[(yr, mo)] = (vc_end / vc_prev - 1) if vc_prev else None
        return out

    icp_ret  = monthly_rets(icp_vc)
    fip_ret  = monthly_rets(fip_vc)
    comp_ret = monthly_rets(comp_vc)

    # ── Step 3: Group into {year: {label: {months, total}}} ───────────────────
    all_yrs = sorted({ym[0] for ym in list(icp_vc) + list(fip_vc) + list(comp_vc)})
    # Get fund name from rentabilidad sheet - use the MOST RECENT year's label
    fund_label = 'FIP'
    try:
        wb2 = openpyxl.load_workbook(path, data_only=True)
        if 'rentabilidad' in wb2.sheetnames:
            ws2 = wb2['rentabilidad']
            best_yr = 0
            cur_yr2 = None
            for r in range(6, ws2.max_row+1):
                yr2 = ws2.cell(r, 3).value
                if isinstance(yr2, (int, float)): cur_yr2 = int(yr2)
                lbl = ws2.cell(r, 4).value
                if (lbl and isinstance(lbl, str) and 'FIP' in lbl.upper()
                        and cur_yr2 and cur_yr2 > best_yr):
                    fund_label = lbl.strip()
                    best_yr = cur_yr2
        wb2.close()
    except: pass

    out = {}
    for yr in all_yrs:
        yr_data = {}
        for label, ret_dict in [('ICP', icp_ret), ('Competencia', comp_ret), (fund_label, fip_ret)]:
            if label == 'Competencia' and not comp_vc: continue
            months = []
            for mo in range(1, 13):
                r = ret_dict.get((yr, mo))
                months.append(r)  # None if missing or fund not started

            if not any(v is not None for v in months): continue

            # Compute annualized total: (end/first_in_year - 1) / n * 12
            vc_dict = icp_vc if label == 'ICP' else (comp_vc if label == 'Competencia' else fip_vc)
            non_none_rets = [v for v in months if v is not None]
            n = len(non_none_rets)
            total = None
            if n > 0:
                # Template formula: (end_vc / first_vc_of_year - 1) / COUNTA * 12
                # COUNTA = 12 for ALL years because return row cells always have formulas
                # (even Jan-Mar = 0 formula for funds that started mid-year)
                # Exception: current partial year uses actual n (months with returns)
                yr_vcs = [(mo, vc_dict.get((yr, mo))) for mo in range(1, 13)
                          if vc_dict.get((yr, mo)) is not None]
                if len(yr_vcs) >= 2:
                    vc_first = yr_vcs[0][1]   # First VC of this year (may be base month)
                    vc_last  = yr_vcs[-1][1]  # Last VC of this year
                    # n_count: 12 for past years (COUNTA counts formula cells as non-empty)
                    #          n for current year (only actual months matter)
                    is_current_yr = (yr == max(all_yrs))
                    n_count = n if is_current_yr else 12
                    total = (vc_last / vc_first - 1) / n_count * 12

            yr_data[label] = {'months': months, 'total': total}
        if yr_data:
            out[yr] = yr_data

    _HIST_CACHE[tmpl_file] = out
    return out

def _get_tmpl_summary(tmpl_file, is_usd):
    """
    For USD funds: read summary directly from template rentabilidad rows.
    Returns (acum_label, icp_row, comp_row, fip_row) as dicts.
    """
    import openpyxl
    path = _TMPL_DIR / tmpl_file
    if not path.exists(): return None, None, None, None
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'rentabilidad' not in wb.sheetnames:
        wb.close(); return None, None, None, None
    ws = wb['rentabilidad']
    acum_label = str(ws.cell(1,24).value or '').replace('\n',' ').strip()

    def read_row(r):
        return {k: ws.cell(r,c).value
                for k,c in zip(['nombre','m','t','s','a','ac'],[19,20,21,22,23,24])}

    if is_usd:
        icp_row  = read_row(2)  # In USD templates R2=Comp (ICP-like)
        fip_row  = read_row(3)  # R3=FIP
        comp_row = read_row(2)  # Comp = same as ICP row for USD
        icp_row['nombre']  = 'ICP (Benchmark)'
        comp_row['nombre'] = 'Competencia'
    else:
        icp_row  = read_row(2)
        comp_row = read_row(3)
        fip_row  = read_row(4)
    wb.close()
    return acum_label, icp_row, comp_row, fip_row

def _fip_row_from_hist(hist_yr, hint):
    """Find the FIP entry in a template year dict. Returns {'months':[], 'total':v}."""
    skip = {'icp','competencia','icp (benchmark)'}
    candidates = [(k,v) for k,v in hist_yr.items() if k.lower() not in skip and 'icp' not in k.lower()]
    if not candidates: return None
    if len(candidates) == 1: return candidates[0][1]
    hint_up = hint.upper().replace('FIP','').replace('VANTRUST','').replace('LIQUIDEZ','').strip()
    for k,v in candidates:
        if hint_up in k.upper().replace('FIP','').replace('LIQUIDEZ','').strip():
            return v
    return candidates[0][1]

def _icp_row_from_hist(hist_yr):
    """Find the ICP entry in a template year dict."""
    for k,v in hist_yr.items():
        if 'icp' in k.lower(): return v
    return None

def _comp_row_from_hist(hist_yr):
    """Find the Competencia entry in a template year dict."""
    for k,v in hist_yr.items():
        if k.lower() == 'competencia': return v
    return None

def _entry_months(entry):
    """Get months list from a hist entry (dict with 'months' key)."""
    if isinstance(entry, dict): return entry.get('months', [])
    return entry or []

def _entry_total(entry):
    """Get stored total from a hist entry."""
    if isinstance(entry, dict): return entry.get('total')
    return None

# ── Build CLP VC series (ODS + template chain for pre-ODS) ───────────────────
def _build_clp_vc(nombre, tmpl_file):
    vc_ods = _get_ods_vc(nombre)
    if not vc_ods: return {}
    if not tmpl_file: return vc_ods

    hist = _get_tmpl_hist(tmpl_file)
    if not hist: return vc_ods

    earliest = min(vc_ods.keys())
    combined = dict(vc_ods)
    cur_vc   = vc_ods[earliest]

    # Collect pre-ODS monthly returns in reverse order
    pre = sorted(
        ((yr, mo_idx+1, ret)
         for yr, yr_data in hist.items()
         for entry in [_fip_row_from_hist(yr_data, nombre)]
         if entry
         for mo_idx, ret in enumerate(_entry_months(entry))
         if ret is not None and (yr, mo_idx+1) < earliest),
        reverse=True
    )

    for yr, mo, ret in pre:
        cur_vc = cur_vc / (1.0 + ret)
        combined[(yr, mo)] = cur_vc

    # Also add one synthetic prior month so first FIP month can be calculated
    if pre:
        oldest_yr, oldest_mo, oldest_ret = pre[-1]
        py, pm = _prev(oldest_yr, oldest_mo, 1)
        if py not in combined or pm not in [k[1] for k in combined if k[0]==py]:
            combined[(py, pm)] = cur_vc / (1.0 + oldest_ret)

    return combined

# ── Return helpers ────────────────────────────────────────────────────────────
def _simple(vc, y, m, n=1):
    v1 = vc.get((y,m)); v0 = vc.get(_prev(y,m,n))
    return v1/v0-1 if v1 and v0 else None

def _annualized(vc, y, m, n=1):
    v1 = vc.get((y,m)); py,pm = _prev(y,m,n); v0 = vc.get((py,pm))
    if not v1 or not v0: return None
    days = (_eom(y,m) - _eom(py,pm)).days
    return (v1/v0-1)/days*360 if days>0 else None

def _ytd(vc, y, m):
    """Acum YYYY: (end/first_in_year - 1) / n_months * 12."""
    v_end = vc.get((y,m))
    if not v_end: return None
    v_first, n = None, 0
    for mm in range(1, m+1):
        v = vc.get((y,mm))
        if v:
            if v_first is None: v_first = v
            n += 1
    if not v_first or n==0 or v_first==v_end: return None
    return (v_end/v_first-1)/n*12

# ── Main entry point ──────────────────────────────────────────────────────────
def _get_tmpl_last_month(tmpl_file: str) -> tuple:
    """Return (year, month) of the last month with FIP data in the template.
    Reads from Datos ICP G column (raw VC) which has cached numeric values."""
    import openpyxl
    path = _TMPL_DIR / tmpl_file
    if not path.exists(): return (0, 0)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if 'Datos ICP (2)' not in wb.sheetnames:
            wb.close(); return (0, 0)
        ws = wb['Datos ICP (2)']
        last_yr, last_mo = 0, 0
        for r in range(3, ws.max_row + 1):
            date_v = ws.cell(r, 1).value
            g_v    = ws.cell(r, 7).value  # col G = raw FIP VC
            if not (date_v and hasattr(date_v, 'year')): continue
            if g_v and isinstance(g_v, (int, float)) and g_v > 0:
                yr, mo = date_v.year, date_v.month
                if (yr, mo) > (last_yr, last_mo):
                    last_yr, last_mo = yr, mo
        wb.close()
        return (last_yr, last_mo)
    except:
        return (0, 0)


def _ods_has_eom(nombre: str, y: int, m: int) -> bool:
    """Check if ODS has end-of-month data (day >= 28) for a fund/month."""
    sql = (f"SELECT MAX(DAY(FECHA_CIERRE)) last_day "
           f"FROM ODS.VALORES_CUOTA_GPI "
           f"WHERE RTRIM(LTRIM(EMPRESA))=N'{nombre}' AND VALOR_CUOTA>0 "
           f"AND YEAR(FECHA_CIERRE)={y} AND MONTH(FECHA_CIERRE)={m}")
    try:
        import requests as _req
        r = _req.post(_ODS_API, json={"Sql": sql},
                      headers={"Content-Type": "application/json"}, timeout=15)
        rows = r.json().get("rows", [{}])
        last_day = rows[0].get("last_day")
        return bool(last_day and int(last_day) >= 28)
    except:
        return False


def _build_clp_vc_from_template(tmpl_file: str) -> dict:
    """Build {(yr,mo): vc} from template's Datos ICP G column (raw FIP VC values)."""
    import openpyxl
    path = _TMPL_DIR / tmpl_file
    if not path.exists(): return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if 'Datos ICP (2)' not in wb.sheetnames:
            wb.close(); return {}
        ws = wb['Datos ICP (2)']
        vc = {}
        for r in range(3, ws.max_row + 1):
            date_v = ws.cell(r, 1).value
            g_v    = ws.cell(r, 7).value
            if date_v and hasattr(date_v, 'year') and g_v and isinstance(g_v, (int, float)) and g_v > 0:
                vc[(date_v.year, date_v.month)] = float(g_v)
        wb.close()
        return vc
    except:
        return {}


def leer_datos_template(nombre_fondo, target_year=None, target_month=None):
    if target_year is None:
        tm = os.environ.get("TARGET_MONTH","").strip()
        if tm: target_year, target_month = map(int, tm.split("-"))
        else:
            hoy = date.today()
            prev = date(hoy.year, hoy.month, 1) - relativedelta(months=1)
            target_year, target_month = prev.year, prev.month
    y, m = target_year, target_month

    display   = NOMBRE_DISPLAY.get(nombre_fondo, nombre_fondo.replace('FIP VANTRUST ',''))
    is_usd    = nombre_fondo in FONDOS_USD
    tmpl_file = FUND_TEMPLATE_MAP.get(nombre_fondo)
    icp       = _get_icp_series()
    vc_comp   = _comp_clp_con_cmf()  # CLP Santander MM: auto CMF + cache JSON

    # ── For USD: read summary and historico directly from template ───────────
    # USD funds: use template for both summary and historico
    # CLP funds: use ODS for summary when available; use template G-column VC when template is current
    if is_usd and tmpl_file:
        return _build_usd_output(nombre_fondo, display, tmpl_file, icp, vc_comp, y, m,
                                  is_usd=True)

    # ── For CLP: ODS + template chain ────────────────────────────────────────
    vc_fip  = _build_clp_vc(nombre_fondo, tmpl_file)

    # If ODS doesn't have END-OF-MONTH data for this month, use template G-column VC
    # ODS may have partial-month data (e.g. day 5) which gives wrong returns
    ods_eom_ok = _ods_has_eom(nombre_fondo, y, m)
    if tmpl_file and not ods_eom_ok:
        tmpl_last = _get_tmpl_last_month(tmpl_file)
        if tmpl_last == (y, m):
            tmpl_vc = _build_clp_vc_from_template(tmpl_file)
            if tmpl_vc and tmpl_vc.get((y, m)):
                vc_fip = tmpl_vc

    # Retorno total del FIP: nivel = VC + dividendos historicos (igual que datos_manager).
    # ICP y Competencia NO llevan dividendo. Sin esto el fondo sale en crudo (~0.65 vs 0.62).
    try:
        from calculos.rentabilidades import DIVIDENDOS
        _div = DIVIDENDOS.get(nombre_fondo, 0.0)
    except Exception:
        _div = 0.0
    if _div:
        vc_fip = {k: v + _div for k, v in vc_fip.items()}

    has_12  = bool(vc_fip.get(_prev(y, m, 12)))

    def calc(vc, es_icp, es_comp, es_fip, name):
        return {
            'nombre': name, 'es_icp': es_icp, 'es_comp': es_comp, 'es_fip': es_fip,
            'm':  _simple(vc, y, m, 1),   't': _simple(vc, y, m, 3),
            's':  _simple(vc, y, m, 6),   'a': _simple(vc, y, m, 12) if has_12 else None,
            'ac': _ytd(vc, y, m),
        }

    resumen = [
        calc(icp,     True,  False, False, 'ICP (Benchmark)'),
        calc(vc_comp, False, True,  False, 'Competencia'),
        calc(vc_fip,  False, False, True,  display),
    ]

    # Acum label from template
    acum_label = f"Acum. {y} (*)"
    if tmpl_file:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_TMPL_DIR/tmpl_file, data_only=True)
            if 'rentabilidad' in wb.sheetnames:
                raw = wb['rentabilidad'].cell(1,24).value
                if raw: acum_label = str(raw).replace('\n',' ').strip()
            wb.close()
        except: pass

    # ── Historical table ─────────────────────────────────────────────────────
    tmpl_hist = _get_tmpl_hist(tmpl_file) if tmpl_file else {}
    historico = _build_clp_historico(y, m, icp, vc_comp, vc_fip, display, tmpl_hist, nombre_fondo)

    # ── Chart ────────────────────────────────────────────────────────────────
    labels, c_icp, c_comp, c_fip = _build_chart(y, m, icp, vc_comp, _get_ods_vc(nombre_fondo))

    return {'nombre_fip': display, 'acum_label': acum_label,
            'resumen': resumen, 'historico': historico,
            'grafico': {'labels': labels, 'icp': c_icp, 'comp': c_comp, 'fip': c_fip}}


def _build_clp_historico(y, m, icp, vc_comp, vc_fip, display, tmpl_hist, nombre_fondo):
    """Build historical table rows for CLP fund.
    Total rule: use template stored total when data is pre-ODS; otherwise _ytd.
    For years entirely from ODS (2026): always use _ytd.
    For years with template data (2024, 2025): use template total directly.
    """
    ODS_START_YR = 2025   # ODS data starts Jul 2025
    out = []
    for yr in [y-2, y-1, y]:
        last_m  = m if yr==y else 12
        yr_tmpl = tmpl_hist.get(yr, {}) if tmpl_hist else {}
        filas   = []

        for lbl, vc, is_icp_like in [('ICP (Benchmark)', icp, True),
                                      ('Competencia', vc_comp, False),
                                      (display, vc_fip, False)]:
            months_out = []
            for mm in range(1, 13):
                if mm > last_m:
                    months_out.append(None); continue

                # Try ODS/extended series first
                has_ods = (vc.get((yr,mm)) and
                           vc.get((yr,mm-1) if mm>1 else _prev(yr,mm)))
                if has_ods:
                    months_out.append(_simple(vc, yr, mm))
                elif not is_icp_like and yr_tmpl:
                    entry = _fip_row_from_hist(yr_tmpl, nombre_fondo)
                    mons  = _entry_months(entry)
                    months_out.append(mons[mm-1] if mons and mm <= len(mons) else None)
                elif is_icp_like and yr_tmpl:
                    # ICP/Comp from template for pre-ODS years
                    if lbl == 'ICP (Benchmark)':
                        entry = _icp_row_from_hist(yr_tmpl)
                    else:
                        entry = _comp_row_from_hist(yr_tmpl)
                    mons = _entry_months(entry) if entry else []
                    val  = mons[mm-1] if mons and mm <= len(mons) else _simple(vc, yr, mm)
                    months_out.append(val)
                else:
                    months_out.append(_simple(vc, yr, mm))

            if not any(v is not None for v in months_out):
                continue

            # Determine total using template stored values (correct formula from Excel)
            # _get_tmpl_hist now computes totals correctly from G/K column data
            total = None
            if yr_tmpl:
                if lbl == 'ICP (Benchmark)':
                    entry = _icp_row_from_hist(yr_tmpl)
                elif lbl == 'Competencia':
                    entry = _comp_row_from_hist(yr_tmpl)
                else:
                    entry = _fip_row_from_hist(yr_tmpl, nombre_fondo)
                total = _entry_total(entry) if entry else None

            if total is None:
                # Fallback for years not in template: compound/n*12
                non_none = [v for v in months_out[:last_m] if v is not None]
                if non_none:
                    n = len(non_none)
                    compound = 1.0
                    for r in non_none: compound *= (1+r)
                    total = (compound - 1) / n * 12

            filas.append({'nombre': lbl, 'meses': months_out, 'total': total})
        if filas:
            out.append({'año': yr, 'filas': filas})
    return out


def _build_usd_output(nombre_fondo, display, tmpl_file, icp, vc_comp, y, m, is_usd=True, use_tmpl_hist=True):
    """For USD funds: read everything from template (more accurate than ODS for these)."""
    import openpyxl
    path = _TMPL_DIR / tmpl_file
    wb   = openpyxl.load_workbook(path, data_only=True)
    ws   = wb['rentabilidad']

    acum_label = str(ws.cell(1,24).value or f'Acum. {y} (*)').replace('\n',' ').strip()

    # Resumen USD: COMPUTAR para el mes objetivo desde los datos (col G + dividendo).
    # M/T/S/A anualizados (/dias*360); Acum = (H/H_ene-1)/n*12. No leer las celdas de
    # formula del resumen (quedan None tras actualizar el template con openpyxl).
    wsd = wb['Datos ICP (2)']
    g_vc, k_vc = {}, {}
    for rr in range(3, wsd.max_row+1):
        fdate = wsd.cell(rr,6).value
        if not (fdate and hasattr(fdate,'year')): continue
        ym = (fdate.year, fdate.month)
        gv = wsd.cell(rr,7).value; kv = wsd.cell(rr,11).value
        if isinstance(gv,(int,float)): g_vc[ym] = float(gv)
        if isinstance(kv,(int,float)): k_vc[ym] = float(kv)
    try:
        from calculos.rentabilidades import DIVIDENDOS
        _div = DIVIDENDOS.get(nombre_fondo, 0.0)
    except Exception:
        _div = 0.0
    h_vc = {ym: v + _div for ym, v in g_vc.items()}   # nivel H = VC + dividendo
    fip_m  = {'m': _annualized(h_vc,y,m,1), 't': _annualized(h_vc,y,m,3),
              's': _annualized(h_vc,y,m,6), 'a': _annualized(h_vc,y,m,12), 'ac': _ytd(h_vc,y,m)}
    comp_m = {'m': _annualized(k_vc,y,m,1), 't': _annualized(k_vc,y,m,3),
              's': _annualized(k_vc,y,m,6), 'a': _annualized(k_vc,y,m,12), 'ac': _ytd(k_vc,y,m)}
    icp_row  = {'nombre': 'ICP (Benchmark)', 'es_icp': True,  'es_comp': False, 'es_fip': False, **comp_m}
    comp_row = {'nombre': 'Competencia',     'es_icp': False, 'es_comp': True,  'es_fip': False, **comp_m}
    fip_row  = {'nombre': display,            'es_icp': False, 'es_comp': False, 'es_fip': True,  **fip_m}

    # Historical from template
    hist_raw = _get_tmpl_hist(tmpl_file)
    wb.close()

    historico = []
    for yr in [y-2, y-1, y]:
        last_m  = m if yr==y else 12
        yr_data = hist_raw.get(yr, {})
        filas   = []
        for lbl, is_icp_like in [('ICP (Benchmark)', True), ('Competencia', False), (display, False)]:
            # Find the right row in template hist
            row_months = None
            if is_icp_like:
                icp_entry = _icp_row_from_hist(yr_data)
                row_months = _entry_months(icp_entry) if icp_entry else None
                entry = icp_entry
                if row_months is None:
                    # Use ICP series for historical ICP
                    row_months = [_simple(icp, yr, mm) for mm in range(1,13)]
                    entry = None
            else:
                comp_rows = [(k,v) for k,v in yr_data.items() if k.lower() == 'competencia']
                if lbl == 'Competencia':
                    entry = comp_rows[0][1] if comp_rows else None
                else:
                    entry = _fip_row_from_hist(yr_data, nombre_fondo)
                row_months = _entry_months(entry)

            if row_months:
                months_out = [row_months[mm-1] if mm<=last_m and mm<=len(row_months) else None
                              for mm in range(1,13)]
                if any(v is not None for v in months_out):
                    # Use stored template total, else compute compound
                    total = _entry_total(entry) if 'entry' in dir() and entry else None
                    if total is None:
                        non_none = [v for v in months_out[:last_m] if v is not None]
                        if non_none:
                            compound = 1.0
                            for r in non_none: compound *= (1+r)
                            n = len(non_none)
                            total = (compound - 1) / n * 12
                    filas.append({'nombre': lbl, 'meses': months_out, 'total': total})
        if filas:
            historico.append({'año': yr, 'filas': filas})

    # Chart from ODS
    vc_ods = _get_ods_vc(nombre_fondo)
    labels, c_icp, c_comp, c_fip = _build_chart(y, m, icp, vc_comp, vc_ods)

    return {'nombre_fip': display, 'acum_label': acum_label,
            'resumen': [icp_row, comp_row, fip_row],
            'historico': historico,
            'grafico': {'labels': labels, 'icp': c_icp, 'comp': c_comp, 'fip': c_fip}}


def _build_chart(y, m, icp, vc_comp, vc_ods):
    start = (y-2, 1); end = (y, m)
    months = sorted({k for k in list(vc_ods.keys())+list(icp.keys())+list(vc_comp.keys())
                     if start <= k <= end})
    if not months: return [], [], [], []
    base = months[0]
    b_f = next((v for k,v in sorted(vc_ods.items()) if k>=base), 1) or 1
    b_i = icp.get(base) or next((v for k,v in sorted(icp.items()) if k>=base), 1)
    b_c = vc_comp.get(base) or next((v for k,v in sorted(vc_comp.items()) if k>=base), 1)
    labels,ci,cc,cf = [],[],[],[]
    for k in months:
        vi = icp.get(k)
        if not vi: continue
        labels.append(f"{k[0]}-{k[1]:02d}")
        ci.append(round(vi/b_i*100, 4))
        vf = vc_ods.get(k); vc_c = vc_comp.get(k)
        cf.append(round(vf/b_f*100,4) if vf else None)
        cc.append(round(vc_c/b_c*100,4) if vc_c else None)
    return labels, ci, cc, cf
