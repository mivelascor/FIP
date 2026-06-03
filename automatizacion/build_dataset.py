#!/usr/bin/env python3
"""
build_dataset.py — Motor de datos para los folletos FIP Vantrust (versión final).

PRINCIPIO RECTOR: los templates son la fuente histórica INTOCABLE. La historia
(fondo / ICP / competencia) se LEE tal cual de la hoja `rentabilidad` de cada
template. Solo se CALCULA y agrega el MES NUEVO. Nunca se recalcula lo viejo.

Los tres insumos automáticos (sin intervención manual):
  1. ICP        -> API del BCCh (serie TIB promedio) + fórmula de composición.
  2. Fondo      -> retorno ajustado por dividendo (VC+d)/(VC_prev+d)-1, con d
                   recuperado automáticamente de la propia historia del template.
  3. Competencia-> scraping CMF (sin captcha). [implementar scrape del formulario]

Fórmulas (validadas al punto base contra VC_para_fichas, mayo 2026):
  Mensual    = nivel_fin / nivel_mes_anterior - 1
  Trimestral = nivel_fin / nivel_hace_3_meses - 1
  Semestral  = nivel_fin / nivel_hace_6_meses - 1
  Anual      = nivel_fin / nivel_hace_12_meses - 1
  Acum (*)   = (nivel_fin / nivel_enero_del_año - 1) / n_meses * 12   [anualizada]
  donde, para el FONDO,  nivel = VC + d (dividendo); para ICP y competencia,
  nivel = el índice/VC directo (sin dividendo).
"""
import openpyxl, datetime, json, os, sys, statistics, urllib.request, urllib.parse
from collections import defaultdict

# ── Config BCCh (credenciales por entorno; NO hardcodear) ───────────────────
BCCH_API  = "https://si3.bcentral.cl/SieteRestWS/SieteRestWS.ashx"
SERIE_TIB = "F022.TIB.TIP.D001.NO.Z.D"   # Tasa promedio transada interbancaria (%)

# ── Fondos activos CLP (los que llegan a abril/mayo). Los USD requieren FX. ──
FUNDS_CLP = {
 "TEMPLATE FONDO LIQUIDEZ ACTIVA.xlsx":      "FIP VANTRUST LIQUIDEZ ACTIVA",
 "TEMPLATE FONDO ALTO APORTE.xlsx":          "FIP VANTRUST LIQUIDEZ ALTO APORTE",
 "TEMPLATE FONDO ALTO CAPITAL.xlsx":         "FIP VANTRUST LIQUIDEZ ALTO CAPITAL",
 "TEMPLATE FONDO LIQUIDEZ ALTO MONTO.xlsx":  "FIP VANTRUST LIQUIDEZ ALTO MONTO",
 "TEMPLATE FONDO LIQUIDEZ CAJA.xlsx":        "FIP VANTRUST LIQUIDEZ CAJA",
 "TEMPLATE FONDO LIQUIDEZ CONTINUA.xlsx":    "FIP VANTRUST LIQUIDEZ CONTINUA",
 "TEMPLATE FONDO LIQUIDEZ CORRIENTE.xlsx":   "FIP VANTRUST LIQUIDEZ CORRIENTE",
 "TEMPLATE FONDO LIQUIDEZ CORTO PLAZO.xlsx": "FIP VANTRUST LIQUIDEZ CORTO PLAZO",
 "TEMPLATE FONDO LIQUIDEZ Disponible I.xlsx":"FIP VANTRUST LIQUIDEZ DISPONIBLE I",
 "TEMPLATE FONDO LIQUIDEZ EFECTIVO.xlsx":    "FIP VANTRUST LIQUIDEZ EFECTIVO",
 "TEMPLATE FONDO LIQUIDEZ FLEXIBLE.xlsx":    "FIP VANTRUST LIQUIDEZ FLEXIBLE",
 "TEMPLATE FONDO LIQUIDEZ UNO.xlsx":         "FIP VANTRUST LIQUIDEZ I",
 "TEMPLATE FONDO LIQUIDEZ LOCAL.xlsx":       "FIP VANTRUST LIQUIDEZ LOCAL",
 "TEMPLATE FONDO LIQUIDEZ Monetario I.xlsx": "FIP VANTRUST LIQUIDEZ MONETARIO I",
 "TEMPLATE FONDO LIQUIDEZ Permanente.xlsx":  "FIP VANTRUST LIQUIDEZ PERMANENTE",
 "TEMPLATE FONDO LIQUIDEZ PLUS.xlsx":        "FIP VANTRUST LIQUIDEZ PLUS",
 "TEMPLATE FONDO LIQUIDEZ Presente.xlsx":    "FIP VANTRUST LIQUIDEZ PRESENTE",
 "TEMPLATE FONDO LIQUIDEZ RENDIMIENTO.xlsx": "FIP VANTRUST LIQUIDEZ RENDIMIENTO",
 "TEMPLATE FONDO LIQUIDEZ SENCILLO.xlsx":    "FIP VANTRUST LIQUIDEZ SENCILLO",
}
# USD: necesitan VC(USD) x tipo de cambio para el retorno en CLP del folleto.
FUNDS_USD = {
 "TEMPLATE FONDO LIQUIDEZ RESERVA DOLAR.xlsx":"FIP VANTRUST LIQUIDEZ RESERVA DOLAR",
 "TEMPLATE FONDO LIQUIDEZ DOLAR.xlsx":        "FIP VANTRUST LIQUIDEZ DOLAR",
 "TEMPLATE FONDO LIQUIDEZ DOLAR CAJA.xlsx":   "FIP VANTRUST LIQUIDEZ DOLAR CAJA",
}

def shift(ym, k=1):
    y, m = map(int, ym.split('-')); m -= k
    while m <= 0: m += 12; y -= 1
    return f"{y}-{m:02d}"

# ── 1. valor_cuota.xlsx -> VC fin de mes por fondo ──────────────────────────
def load_valor_cuota(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Consulta1 (2)"] if "Consulta1 (2)" in wb.sheetnames else wb[wb.sheetnames[0]]
    eom, ld = defaultdict(dict), defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        f, n, p = row[0], row[1], row[4]
        if n is None or not isinstance(f, datetime.datetime): continue
        nm = str(n).strip().strip("'"); ym = f"{f.year}-{f.month:02d}"
        if ym not in ld[nm] or f.day > ld[nm][ym]:
            ld[nm][ym] = f.day; eom[nm][ym] = float(p)
    return eom

# ── 2. ICP desde el BCCh (exacto). Base = último ICP del template. ──────────
def icp_eom_bcch(icp_base, base_ym, target_ym):
    """Compone la TIB diaria del BCCh desde fin de base_ym hasta fin de target_ym.
    ICP_i = ICP_{i-1} * (1 + TIB_{i-1}/100 * Ndias/360). Devuelve nivel ICP EOM."""
    user = os.environ.get("BCCH_USER", ""); pw = os.environ.get("BCCH_PASS", "")
    if not user or not pw:
        raise EnvironmentError("Faltan BCCH_USER / BCCH_PASS en el entorno.")
    by, bm = map(int, base_ym.split('-')); ty, tm = map(int, target_ym.split('-'))
    d_ini = datetime.date(by, bm, 1) + datetime.timedelta(days=31)
    d_ini = (d_ini.replace(day=1) - datetime.timedelta(days=1))  # fin base_ym
    d_fin = (datetime.date(ty + (tm == 12), (tm % 12) + 1, 1) - datetime.timedelta(days=1))
    p = {"user": user, "pass": pw, "function": "GetSeries", "timeseries": SERIE_TIB,
         "firstdate": d_ini.strftime("%Y-%m-%d"), "lastdate": d_fin.strftime("%Y-%m-%d")}
    raw = urllib.request.urlopen(urllib.request.Request(
        BCCH_API + "?" + urllib.parse.urlencode(p), headers={"User-Agent": "Mozilla/5.0"}),
        timeout=60).read().decode("utf-8", "replace")
    obs = (json.loads(raw).get("Series") or {}).get("Obs") or []
    tib = {datetime.datetime.strptime(o["indexDateString"], "%d-%m-%Y").date(): float(o["value"])
           for o in obs if o["statusCode"] == "OK"}
    icp = float(icp_base); last = tib.get(d_ini) or next(iter(sorted(tib.values())), 4.5)
    day = d_ini
    while day < d_fin:
        t = tib.get(day, last); last = t
        icp *= (1 + t/100 * 1/360)
        day += datetime.timedelta(days=1)
    return icp

# ── 3. Lee historia oficial (rentabilidad) y nivel ICP/comp del template ────
def read_template_history(tmpl):
    """Devuelve (icp_levels{ym:val}, comp_levels{ym:val}, fip_returns{ym:r})."""
    wb = openpyxl.load_workbook(tmpl, data_only=True)
    icp_lv, comp_lv = {}, {}
    ws = wb["Datos ICP (2)"]
    for row in ws.iter_rows():
        a = row[0].value
        if not isinstance(a, datetime.datetime): continue
        ym = f"{a.year}-{a.month:02d}"
        b = row[1].value if len(row) > 1 else None
        k = row[10].value if len(row) > 10 else None
        if isinstance(b, (int, float)): icp_lv[ym] = float(b)
        if isinstance(k, (int, float)): comp_lv[ym] = float(k)
    fip = {}
    if "rentabilidad" in wb.sheetnames:
        wr = wb["rentabilidad"]; cy = None
        for r in range(1, wr.max_row + 1):
            c = wr.cell(r, 3).value; d = wr.cell(r, 4).value
            if isinstance(c, (int, float)) and 2000 < c < 2100: cy = int(c)
            if d and isinstance(d, str):
                du = d.strip().upper()
                isfip = du != "ICP" and "COMPETENCIA" not in du and \
                        ("FIP" in du or "LIQUIDEZ" in du or "ALTO" in du)
                if isfip and cy:
                    for mi in range(12):
                        v = wr.cell(r, 5 + mi).value
                        if isinstance(v, (int, float)) and v != 0:
                            fip[f"{cy}-{mi+1:02d}"] = float(v)
    return icp_lv, comp_lv, fip

def recover_dividend(vc, fip_ret):
    """Recupera d del mes oficial MÁS RECIENTE y lo valida contra el mes previo.
    Si el historial no es consistente con (VC+d) (p.ej. fondos cuyo retorno oficial
    NO sale del VC del query), devuelve 0.0 -> se usa el VC crudo del query."""
    oms = [m for m in sorted(fip_ret) if m in vc and shift(m) in vc]
    if not oms: return 0.0
    m = oms[-1]; r = fip_ret[m]
    if r == 0: return 0.0
    d = (vc[m] - (1 + r) * vc[shift(m)]) / r
    # validación: ¿d reproduce el mes previo? si no, el historial no viene del VC.
    if len(oms) >= 2:
        p = oms[-2]
        calc = (vc[p] + d) / (vc[shift(p)] + d) - 1
        if abs(calc - fip_ret[p]) > 0.0005:
            return 0.0
    return d

# ── 4. Métricas de período sobre una serie de niveles ──────────────────────
def metrics(levels, end):
    def ret(e, a):
        return None if (e not in levels or a not in levels or not levels[a]) \
               else round((levels[e]/levels[a] - 1) * 100, 2)
    y = end.split('-')[0]; jan = f"{y}-01"
    months = [k for k in levels if k.startswith(y) and k <= end]
    acum = None
    if jan in levels and levels[jan] and months:
        acum = round((levels[end]/levels[jan] - 1)/len(months) * 12 * 100, 2)
    return {"mensual": ret(end, shift(end, 1)), "trimestral": ret(end, shift(end, 3)),
            "semestral": ret(end, shift(end, 6)), "anual": ret(end, shift(end, 12)),
            "acum": acum}

# ── 5. Orquestación: agrega SOLO el mes nuevo, lee historia del template ────
def build(valor_cuota_path, templates_dir, end):
    eom = load_valor_cuota(valor_cuota_path)
    dataset = {}
    todos = [(t, n, "CLP") for t, n in FUNDS_CLP.items()] + \
            [(t, n, "USD") for t, n in FUNDS_USD.items()]
    for tmpl, nemo, moneda in todos:
        path = os.path.join(templates_dir, tmpl)
        if not os.path.exists(path): continue
        vc = eom.get(nemo, {})
        if end not in vc:                       # inactivo o sin VC del mes
            continue
        icp_lv, comp_lv, fip_ret = read_template_history(path)
        base_ym = max(icp_lv)                    # último ICP histórico del template
        # nivel ICP del mes nuevo desde BCCh (si falla, deja el último histórico)
        try:
            icp_lv[end] = icp_eom_bcch(icp_lv[base_ym], base_ym, end)
        except Exception as e:
            icp_lv[end] = icp_lv[base_ym]
        # fondo: nivel H = VC + d
        d = recover_dividend(vc, fip_ret)
        H = {m: v + d for m, v in vc.items()}
        # competencia del mes nuevo: CMF (Santander CLP / Banchile USD)
        try:
            from competencia_cmf import valor_cuota_competencia
            comp_lv[end] = valor_cuota_competencia(moneda, end)
        except Exception:
            if comp_lv and end not in comp_lv:
                comp_lv[end] = comp_lv[max(comp_lv)]
        dataset[nemo] = {
            "template": tmpl, "moneda": moneda, "mes": end,
            "dividendo": round(d, 4), "vc": round(vc[end], 4),
            "fondo": metrics(H, end),
            "icp": metrics(icp_lv, end),
            "competencia": metrics(comp_lv, end) if comp_lv else None,
        }
    return dataset

if __name__ == "__main__":
    vc_path = sys.argv[1] if len(sys.argv) > 1 else "inputs/valor_cuota.xlsx"
    tdir    = sys.argv[2] if len(sys.argv) > 2 else "inputs/templates"
    end     = sys.argv[3] if len(sys.argv) > 3 else datetime.date.today().strftime("%Y-%m")
    ds = build(vc_path, tdir, end)
    print(f"[{end}] {len(ds)} fondos CLP procesados")
    for nemo, d in ds.items():
        f = d["fondo"]
        print(f"  {nemo.replace('FIP VANTRUST LIQUIDEZ ',''):<16} "
              f"M={f['mensual']} T={f['trimestral']} S={f['semestral']} "
              f"A={f['anual']} Acum={f['acum']}  (div={d['dividendo']})")
